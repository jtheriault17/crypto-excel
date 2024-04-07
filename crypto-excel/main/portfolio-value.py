from openpyxl import load_workbook
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font
import json
import os.path

def load_portfolio_value():
    portfolio_value = {}
    if os.path.exists('../crypto-excel/data/portfolio-value.json'):
        with open('../crypto-excel/data/portfolio-value.json', 'r') as f:
            try:
                portfolio_value = json.load(f)
            except json.JSONDecodeError:
                print("Error loading portfolio values. Initializing empty portfolio values.")
    return portfolio_value

def load_portfolio():
    portfolio = {}
    if os.path.exists('../crypto-excel/data/portfolio.json'):
        with open('../crypto-excel/data/portfolio.json', 'r') as f:
            try:
                portfolio = json.load(f)
            except json.JSONDecodeError:
                print("Error loading portfolio. Initializing empty portfolio.")
    return portfolio

def load_cost_basis():
    cost_basis = {}
    if os.path.exists('../crypto-excel/data/cost_basis.json'):
        with open('../crypto-excel/data/cost_basis.json', 'r') as f:
            try:
                cost_basis = json.load(f)
            except json.JSONDecodeError:
                print("Error loading cost basis. Initializing empty cost basis.")
    return cost_basis

def load_coin_id_dict():
    coin_id_dict = {}
    if os.path.exists('../crypto-excel/data/coin-id-dictionary.json'):
        with open('../crypto-excel/data/coin-id-dictionary.json', 'r') as f:
            try:
                coin_id_dict = json.load(f)
            except json.JSONDecodeError:
                print("Error loading coin ID dictionary. Initializing empty dictionary.")
    return coin_id_dict

def get_coin_id(symbol):
    coin_id_dict = load_coin_id_dict()
    return coin_id_dict[symbol]

# Load historical data from Excel workbook
def load_historical_data(symbol):
    wb = load_workbook('../crypto-excel/workbooks/historical-data.xlsx', read_only=True, data_only=True)
    ws = wb[get_coin_id(symbol.lower())]
    historical_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            historical_data[row[0].date()] = row[1]
    return historical_data

# Load transactions data from Excel workbook
def load_transaction():
    workbook_path = '../crypto-excel/workbooks/Transactions.xlsm'
    wb = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    ws = wb['Transactions']
    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            transactions.append(row)
    return transactions

# Load total data from Excel workbook
def get_total_data_dates():
    wb = load_workbook(filename='../crypto-excel/workbooks/total-data.xlsx', read_only=False, data_only=True)
    ws = wb['Total Data']
    dates = []

    start_date = datetime.today()
    end_date = start_date - timedelta(days=365)

    row_number = 6

    while start_date >= end_date:
        # Write the date to the cell in column A
        ws.cell(row=row_number, column=1, value=start_date)
        # Append the date to the list
        dates.append(start_date)
        # Move to the next row
        row_number += 1
        # Subtract one day from the start date
        start_date -= timedelta(days=1)

    wb.save(filename='../crypto-excel/workbooks/total-data.xlsx')
    return dates

def get_cost_basis(transactions, dates):
    cost_basis = load_cost_basis()
    for date in dates:
        date_str = date.strftime('%m/%d/%y')

def get_portfolio(transactions, dates):
    portfolio = load_portfolio()
    for date in dates:
        date_str = date.strftime('%m/%d/%y')
        if date_str not in portfolio:
            symbol_quantity = {}
            for transaction_date, type, received_quantity, received_currency, received_cost_basis, sent_quantity, sent_currency, sent_cost_basis, fee_amount, fee_currency, fee_cost_basis, realized_return, fee_realized_return in transactions:
                if transaction_date.date() < date.date():
                    if received_currency not in symbol_quantity:
                        symbol_quantity[received_currency] = 0
                    if sent_currency not in symbol_quantity:
                        symbol_quantity[sent_currency] = 0
                    if fee_currency not in symbol_quantity:
                        symbol_quantity[fee_currency] = 0

                    if received_quantity:
                        symbol_quantity[received_currency] += received_quantity
                    if sent_quantity:
                        symbol_quantity[sent_currency] -= sent_quantity
                    if fee_amount:
                        symbol_quantity[fee_currency] -= fee_amount
            # Remove symbols with quantity <= 0
            symbol_quantity = {symbol: quantity for symbol, quantity in symbol_quantity.items() if quantity > 0}
           # Calculate value for each symbol using calculate_symbol_value() function
            symbol_values = {}
            for symbol, quantity in symbol_quantity.items():
                value = calculate_symbol_value(symbol, quantity, date)
                if value > 0:
                    symbol_values[symbol] = {"quantity": quantity, "value": value}
            if symbol_values != None:
                portfolio[date_str] = symbol_values
    with open('../crypto-excel/data/portfolio.json', 'w') as f:
                json.dump(portfolio, f, indent=4)
    return portfolio

# Calculate portfolio value for a given date
def calculate_portfolio_value(date, portfolio):
    portfolio_value = 0
    for symbol, data in portfolio[date].items():
        portfolio_value += data.get('value', 0)
    return portfolio_value

def calculate_symbol_value(symbol, quantity, date):
    historical_data = load_historical_data(symbol)
    if date.date() in historical_data:
        value = quantity * historical_data[date.date()]
        return value
    return 0

def get_portfolio_values(dates, portfolio):
    portfolio_values = load_portfolio_value()
    for date in dates:
        date_str = date.strftime('%m/%d/%y')
        if date_str not in portfolio_values:
            portfolio_values[date_str] = calculate_portfolio_value(date_str, portfolio)
    
    with open('../crypto-excel/data/portfolio-value.json', 'w') as f:
                json.dump(portfolio_values, f, indent=4)

    return portfolio_values

# Main function to populate 'Total Data' sheet
def write_portfolio_values(portfolio_values):
    wb = load_workbook(filename='../crypto-excel/workbooks/total-data.xlsx', data_only=True)

    ws = wb['Total Data']

    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    row_number = 6
    while True:
        date_cell = ws.cell(row=row_number, column=1)
        date = date_cell.value
        if not date:
            break
        date_str = date.strftime('%m/%d/%y')
        portfolio_value = portfolio_values[date_str]
        if portfolio_value is not None:
            ws.cell(row=row_number, column=3, value=portfolio_value)
            ws.cell(row=row_number, column=3).fill = white_fill
        row_number += 1
        
    wb.save(filename='../crypto-excel/workbooks/total-data.xlsx')

def write_portfolio(portfolio):
    wb = load_workbook(filename='../crypto-excel/workbooks/total-data.xlsx', data_only=True)

    ws = wb['Total Data']

    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    bold = Font(bold=True)
    not_bold = Font(bold=False)

    row_number = 6
    date_row_number = 6

    while True:
        date_cell = ws.cell(row=date_row_number, column=1)
        date = date_cell.value
        if not date:
            break
        date_str = date.strftime('%m/%d/%y')
        
        # Check if the date exists in the portfolio
        if date_str in portfolio:
            symbol_values = portfolio[date_str]
            
            # Write the date at the top of the group of symbols
            date_cell = ws.cell(row=row_number, column=5)
            date_cell.value = date_str
            date_cell.fill = white_fill
            date_cell.font = bold

            ws.cell(row=row_number, column=6).value = None
            ws.cell(row=row_number, column=7).value = None
            ws.cell(row=row_number, column=6).fill = white_fill
            ws.cell(row=row_number, column=7).fill = white_fill
            
            # Increment row_number to move to the next row for symbols
            row_number += 1
            
            # Iterate over symbols and their values for the current date
            for symbol, data in symbol_values.items():
                quantity = data.get('quantity', 0)
                value = data.get('value', 0)

                if value > 1:
                    symbol_cell = ws.cell(row=row_number, column=5)
                    quantity_cell = ws.cell(row=row_number, column=6)
                    value_cell = ws.cell(row=row_number, column=7)

                    symbol_cell.value = symbol
                    quantity_cell.value = quantity
                    value_cell.value = value

                    symbol_cell.fill = white_fill
                    symbol_cell.font = not_bold

                    quantity_cell.fill = white_fill
                    quantity_cell.font = not_bold

                    value_cell.fill = white_fill
                    value_cell.font = not_bold

                    row_number += 1
        
        ws.cell(row=row_number, column=5).value = None
        ws.cell(row=row_number, column=6).value = None
        ws.cell(row=row_number, column=7).value = None
        ws.cell(row=row_number, column=5).fill = white_fill
        ws.cell(row=row_number, column=6).fill = white_fill
        ws.cell(row=row_number, column=7).fill = white_fill

        date_row_number += 1
        row_number += 1
        
    wb.save(filename='../crypto-excel/workbooks/total-data.xlsx')

# Main function
def main():
    transactions = load_transaction()
    dates = get_total_data_dates()

    portfolio = get_portfolio(transactions, dates)
    portfolio_values = get_portfolio_values(dates, portfolio)

    write_portfolio_values(portfolio_values)
    write_portfolio(portfolio)


if __name__ == "__main__":
    main()
