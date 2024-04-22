import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import time
import json
import load

# Function to get coin ID for a given symbol
def get_coin_id(symbol):
    coin_list = load.load_coin_list
    coin_id_dict = load.load_coin_id_dict()
    if symbol in coin_id_dict:
        return coin_id_dict[symbol]
    else:
        matching_coins = [coin for coin in coin_list if coin['symbol'] == symbol]
        if not matching_coins:
            return None
        elif len(matching_coins) == 1:
            coin_id = matching_coins[0]['id']
            coin_id_dict[symbol.lower()] = coin_id
            with open('../crypto-excel/data/coin-id-dictionary.json', 'w') as f:
                json.dump(coin_id_dict, f, indent=4)
            return coin_id
        else:
            print(f"Multiple coins found for symbol '{symbol}':")
            for i, coin in enumerate(matching_coins, 1):
                print(f"{i}. {coin['name']}")
            choice = input("Enter the number corresponding to the desired coin: ")
            while not choice.isdigit() or int(choice) < 1 or int(choice) > len(matching_coins):
                choice = input("Invalid input. Please enter a valid number: ")
            coin_id = matching_coins[int(choice) - 1]['id']
            coin_id_dict[symbol.lower()] = coin_id
            with open('../crypto-excel/data/coin-id-dictionary.json', 'w') as f:
                json.dump(coin_id_dict, f, indent=4)
            return coin_id

def get_symbols():
    excel_filename = "../crypto-excel/workbooks/Transactions.xlsm"
    sheet_name = "Currency Data"

    wb_transactions = load_workbook(filename=excel_filename, read_only=True, data_only=True)
    ws_transactions = wb_transactions[sheet_name]

    excel_filename = "../crypto-excel/workbooks/historical-data.xlsx"
    wb_data = load_workbook(filename=excel_filename, read_only=True, data_only=True)

    symbols = []
    for row in ws_transactions.iter_rows(min_row=2, max_row=ws_transactions.max_row, min_col=1, max_col=8, values_only=True):
        if row[7] and row[7] > 5 or (row[0] and get_coin_id(row[0].lower()) not in wb_data.sheetnames):
            symbols.append(row[0].lower())
    return symbols

def get_all_symbols():
    excel_filename = "../crypto-excel/workbooks/Transactions.xlsm"
    sheet_name = "Currency Data"

    wb = load_workbook(filename=excel_filename, read_only=True, data_only=True)
    ws = wb[sheet_name]

    symbols = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        if row[0]:
            symbols.append(row[0].lower())
    return symbols

def create_coin_id_dict(symbols):
    coin_ids = {}
    for symbol in symbols:
        coin_id = get_coin_id(symbol)
        if coin_id:
            coin_ids[symbol] = coin_id
        else:
            print(f"No matching coin found for symbol '{symbol}'")
    return coin_ids


# Function to fetch data for a given coin and write it to Excel
def fetch_and_write(coin_id, excel_filename):
    url = f"https://api.coingecko.com/api/v3/coins/{coin_id}/market_chart?vs_currency=usd&days=365&interval=daily&precision=8"
    headers = {"x-cg-demo-api-key": "CG-U3VbGJ3KKNE5dVgvttoKb1dv"}
    response = requests.get(url, headers=headers)
    json_data = response.json()

    wb = load_workbook(filename=excel_filename)

    # Check if sheet for the coin already exists, if not, create a new one
    if coin_id in wb.sheetnames:
        ws = wb[coin_id]
        # Clear existing data in the sheet
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=coin_id)

    # Extracting data
    prices = json_data["prices"]
    market_caps = json_data["market_caps"]
    total_volumes = json_data["total_volumes"]

    combined_data = {}
    for price in prices:
        timestamp = datetime.fromtimestamp(price[0] / 1000)
        combined_data[timestamp] = {"Date": timestamp, "Price": price[1]}

    for market_cap in market_caps:
        timestamp = datetime.fromtimestamp(market_cap[0] / 1000)
        if timestamp in combined_data:
            combined_data[timestamp]["Market Caps"] = market_cap[1]

    for total_volume in total_volumes:
        timestamp = datetime.fromtimestamp(total_volume[0] / 1000)
        if timestamp in combined_data:
            combined_data[timestamp]["Total Volumes"] = total_volume[1]
    
    # Writing data to sheet
    ws.append(["Date", "Price", "Market Caps", "Total Volumes"])
    for data in combined_data.values():
        ws.append([data.get("Date"), data.get("Price"), data.get("Market Caps"), data.get("Total Volumes")])

    # Save the workbook
    wb.save(filename=excel_filename)

def write_historical_data(coin_ids):
    excel_filename = "../crypto-excel/workbooks/historical-data.xlsx"
    # Loop through each coin and fetch data
    for coin, coin_id in coin_ids.items():
        print(f"Fetching data for {coin}...")
        retry_count = 0
        while retry_count < 2:  # Retry only once
            try:
                fetch_and_write(coin_id, excel_filename)
                print(f"Data for {coin} fetched and written to Excel.")
                break  # Exit loop if successful
            except Exception as e:
                print(f"Error fetching data for {coin}: {e}")
                retry_count += 1
                time.sleep(30)  # Sleep for 30 seconds before retrying
        else:
            print(f"Failed to fetch data for {coin} after 1 retry.")

def fetch_market_data(coin_ids):
    coin_ids_str = '%2C'.join(coin_ids.values())
    url = f"https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&ids={coin_ids_str}&per_page=100&page=1&sparkline=false&price_change_percentage=1h%2C24h%2C7d%2C14d%2C30d%2C200d%2C1y&locale=en&precision=8"
    headers = {"x-cg-demo-api-key": "CG-U3VbGJ3KKNE5dVgvttoKb1dv"}
    response = requests.get(url, headers=headers)
    json_data = response.json()

    with open('../crypto-excel/data/market-data.json', 'w') as f:
                json.dump(json_data, f, indent=4)
    return json_data


def main():
    current_symbols = get_symbols()
    all_symbols = get_all_symbols()
    current_coin_ids = create_coin_id_dict(current_symbols)
    all_coin_ids = create_coin_id_dict(all_symbols)
    fetch_market_data(all_coin_ids)
    write_historical_data(current_coin_ids)
    

if __name__ == "__main__":
    main()

